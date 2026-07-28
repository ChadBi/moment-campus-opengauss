#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
生成数据库设计产物：
1. Excel 表结构文档（21 张表，每表一个 Sheet + 总览 Sheet）
2. ER 图（SVG 格式，总体 + 5 个子系统）
3. ER 图（DOT 源码，供 Graphviz 渲染）

使用虚拟环境：backend/.venv
运行命令：backend\\.venv\\Scripts\\python.exe backend\\scripts\\generate_db_design.py
输出目录：docs/design/
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# 1. 21 张表的完整定义（基于 backend/app/models/ 实际代码）
# =============================================================================

# 字段定义格式：(字段名, 数据类型, 是否主键, 外键引用表, 是否可空, 默认值, 说明)
TABLES = [
    {
        "name": "schools",
        "cn_name": "学校表",
        "desc": "存储学校信息，本项目仅江南大学蠡湖校区 1 条记录",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("name", "VARCHAR(100)", False, None, False, None, "学校名称"),
            ("code", "VARCHAR(20)", False, None, False, None, "学校代码（唯一）"),
            ("logo_url", "VARCHAR(500)", False, None, True, None, "校徽URL"),
            ("province", "VARCHAR(50)", False, None, True, None, "省份"),
            ("city", "VARCHAR(50)", False, None, True, None, "城市"),
            ("address", "VARCHAR(255)", False, None, True, None, "详细地址"),
            ("center_lat", "NUMERIC(10,7)", False, None, True, None, "中心纬度"),
            ("center_lng", "NUMERIC(10,7)", False, None, True, None, "中心经度"),
            ("map_zoom", "INTEGER", False, None, True, "15", "地图缩放级别"),
            ("is_active", "BOOLEAN", False, None, False, "TRUE", "是否启用"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
            ("updated_at", "TIMESTAMP", False, None, False, "CURRENT", "更新时间"),
        ],
    },
    {
        "name": "users",
        "cn_name": "用户表",
        "desc": "存储系统用户信息，含普通用户、管理员、超级管理员",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("email", "VARCHAR(255)", False, None, False, None, "邮箱（唯一）"),
            ("nickname", "VARCHAR(50)", False, None, False, None, "昵称"),
            ("password_hash", "VARCHAR(255)", False, None, False, None, "密码哈希（bcrypt）"),
            ("avatar_url", "VARCHAR(500)", False, None, True, None, "头像URL"),
            ("school_id", "BIGINT", False, "schools", False, None, "学校ID（外键）"),
            ("role", "VARCHAR(20)", False, None, False, "'user'", "角色：user/admin/super_admin"),
            ("bio", "VARCHAR(500)", False, None, True, None, "个人简介"),
            ("is_active", "BOOLEAN", False, None, False, "TRUE", "是否启用"),
            ("last_login_at", "TIMESTAMP", False, None, True, None, "最后登录时间"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
            ("updated_at", "TIMESTAMP", False, None, False, "CURRENT", "更新时间"),
            ("is_deleted", "BOOLEAN", False, None, False, "FALSE", "软删除标记"),
            ("deleted_at", "TIMESTAMP", False, None, True, None, "删除时间"),
        ],
    },
    {
        "name": "categories",
        "cn_name": "分类表",
        "desc": "信息分类，如活动、失物、求助等",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("name", "VARCHAR(50)", False, None, False, None, "分类名称"),
            ("code", "VARCHAR(30)", False, None, False, None, "分类代码（唯一）"),
            ("icon", "VARCHAR(10)", False, None, False, None, "图标emoji"),
            ("description", "VARCHAR(200)", False, None, True, None, "描述"),
            ("default_validity_days", "INTEGER", False, None, False, "30", "默认有效天数"),
            ("sort_order", "INTEGER", False, None, False, "0", "排序"),
            ("is_active", "BOOLEAN", False, None, False, "TRUE", "是否启用"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
            ("updated_at", "TIMESTAMP", False, None, False, "CURRENT", "更新时间"),
        ],
    },
    {
        "name": "locations",
        "cn_name": "地点表",
        "desc": "校园地点，含经纬度，关联学校",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("school_id", "BIGINT", False, "schools", False, None, "学校ID（外键）"),
            ("name", "VARCHAR(100)", False, None, False, None, "地点名称"),
            ("description", "VARCHAR(500)", False, None, True, None, "描述"),
            ("latitude", "NUMERIC(10,7)", False, None, False, None, "纬度"),
            ("longitude", "NUMERIC(10,7)", False, None, False, None, "经度"),
            ("floor", "VARCHAR(10)", False, None, True, None, "楼层"),
            ("building", "VARCHAR(100)", False, None, True, None, "建筑名"),
            ("post_count", "INTEGER", False, None, False, "0", "关联信息数"),
            ("is_verified", "BOOLEAN", False, None, False, "FALSE", "是否已验证"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
            ("updated_at", "TIMESTAMP", False, None, False, "CURRENT", "更新时间"),
            ("is_deleted", "BOOLEAN", False, None, False, "FALSE", "软删除标记"),
            ("deleted_at", "TIMESTAMP", False, None, True, None, "删除时间"),
        ],
    },
    {
        "name": "posts",
        "cn_name": "信息表",
        "desc": "核心表，存储校园信息，含6态状态机",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("user_id", "BIGINT", False, "users", False, None, "发布者ID（外键）"),
            ("school_id", "BIGINT", False, "schools", False, None, "学校ID（外键）"),
            ("category_id", "BIGINT", False, "categories", False, None, "分类ID（外键）"),
            ("location_id", "BIGINT", False, "locations", True, None, "地点ID（外键，可空）"),
            ("title", "VARCHAR(200)", False, None, False, None, "标题"),
            ("content", "TEXT", False, None, False, None, "内容"),
            ("is_anonymous", "BOOLEAN", False, None, False, "FALSE", "是否匿名"),
            ("status", "VARCHAR(20)", False, None, False, "'pending'", "状态：6态状态机"),
            ("view_count", "INTEGER", False, None, False, "0", "浏览数"),
            ("like_count", "INTEGER", False, None, False, "0", "点赞数"),
            ("comment_count", "INTEGER", False, None, False, "0", "评论数"),
            ("favorite_count", "INTEGER", False, None, False, "0", "收藏数"),
            ("valid_count", "INTEGER", False, None, False, "0", "证实数"),
            ("invalid_count", "INTEGER", False, None, False, "0", "证伪数"),
            ("expire_at", "TIMESTAMP", False, None, True, None, "信息截止时间"),
            ("lost_type", "VARCHAR(10)", False, None, True, None, "失物类型：lost/found"),
            ("contact_info", "VARCHAR(255)", False, None, True, None, "联系方式"),
            ("is_top", "BOOLEAN", False, None, False, "FALSE", "是否置顶"),
            ("is_recommend", "BOOLEAN", False, None, False, "FALSE", "是否推荐"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
            ("updated_at", "TIMESTAMP", False, None, False, "CURRENT", "更新时间"),
            ("is_deleted", "BOOLEAN", False, None, False, "FALSE", "软删除标记"),
            ("deleted_at", "TIMESTAMP", False, None, True, None, "删除时间"),
        ],
    },
    {
        "name": "post_images",
        "cn_name": "信息图片表",
        "desc": "信息的附图，一对多关联 Post",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("post_id", "BIGINT", False, "posts", False, None, "信息ID（外键）"),
            ("image_url", "VARCHAR(500)", False, None, False, None, "图片URL"),
            ("thumbnail_url", "VARCHAR(500)", False, None, True, None, "缩略图URL"),
            ("sort_order", "INTEGER", False, None, False, "0", "排序"),
            ("file_size", "INTEGER", False, None, True, None, "文件大小(字节)"),
            ("width", "INTEGER", False, None, True, None, "宽度(px)"),
            ("height", "INTEGER", False, None, True, None, "高度(px)"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
            ("is_deleted", "BOOLEAN", False, None, False, "FALSE", "软删除标记"),
            ("deleted_at", "TIMESTAMP", False, None, True, None, "删除时间"),
        ],
    },
    {
        "name": "comments",
        "cn_name": "评论表",
        "desc": "信息评论，支持自引用实现评论回复树",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("post_id", "BIGINT", False, "posts", False, None, "信息ID（外键）"),
            ("user_id", "BIGINT", False, "users", False, None, "评论者ID（外键）"),
            ("parent_id", "BIGINT", False, "comments", True, None, "父评论ID（自引用外键）"),
            ("reply_to_user_id", "BIGINT", False, "users", True, None, "回复目标用户ID（外键）"),
            ("content", "TEXT", False, None, False, None, "评论内容"),
            ("like_count", "INTEGER", False, None, False, "0", "点赞数"),
            ("status", "VARCHAR(20)", False, None, False, "'pending'", "状态"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
            ("updated_at", "TIMESTAMP", False, None, False, "CURRENT", "更新时间"),
            ("is_deleted", "BOOLEAN", False, None, False, "FALSE", "软删除标记"),
            ("deleted_at", "TIMESTAMP", False, None, True, None, "删除时间"),
        ],
    },
    {
        "name": "likes",
        "cn_name": "点赞表",
        "desc": "用户对信息的点赞记录，post_id+user_id 联合唯一",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("post_id", "BIGINT", False, "posts", False, None, "信息ID（外键）"),
            ("user_id", "BIGINT", False, "users", False, None, "用户ID（外键）"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
        ],
    },
    {
        "name": "favorites",
        "cn_name": "收藏表",
        "desc": "用户对信息的收藏记录，post_id+user_id 联合唯一",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("post_id", "BIGINT", False, "posts", False, None, "信息ID（外键）"),
            ("user_id", "BIGINT", False, "users", False, None, "用户ID（外键）"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
        ],
    },
    {
        "name": "validation_records",
        "cn_name": "协同验证记录表",
        "desc": "用户对信息的协同验证，含5类：证实/证伪/更新/过期报告/冲突报告",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("post_id", "BIGINT", False, "posts", False, None, "信息ID（外键）"),
            ("user_id", "BIGINT", False, "users", False, None, "验证者ID（外键）"),
            ("validation_type", "VARCHAR(30)", False, None, False, None, "验证类型：5类"),
            ("comment", "VARCHAR(500)", False, None, True, None, "验证评论"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
        ],
    },
    {
        "name": "reports",
        "cn_name": "举报表",
        "desc": "用户举报信息或评论，含双角色引用（举报者+处理者）",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("post_id", "BIGINT", False, "posts", True, None, "信息ID（外键，可空）"),
            ("comment_id", "BIGINT", False, "comments", True, None, "评论ID（外键，可空）"),
            ("reporter_id", "BIGINT", False, "users", False, None, "举报者ID（外键）"),
            ("report_type", "VARCHAR(30)", False, None, False, None, "举报类型"),
            ("description", "TEXT", False, None, True, None, "举报描述"),
            ("status", "VARCHAR(20)", False, None, False, "'pending'", "处理状态"),
            ("handler_id", "BIGINT", False, "users", True, None, "处理者ID（外键，可空）"),
            ("handle_result", "TEXT", False, None, True, None, "处理结果"),
            ("handled_at", "TIMESTAMP", False, None, True, None, "处理时间"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
            ("updated_at", "TIMESTAMP", False, None, False, "CURRENT", "更新时间"),
        ],
    },
    {
        "name": "notifications",
        "cn_name": "通知表",
        "desc": "系统通知，含双角色引用（接收者+触发者），多态关联目标",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("user_id", "BIGINT", False, "users", False, None, "接收者ID（外键）"),
            ("type", "VARCHAR(30)", False, None, False, None, "通知类型"),
            ("title", "VARCHAR(200)", False, None, False, None, "通知标题"),
            ("content", "VARCHAR(500)", False, None, True, None, "通知内容"),
            ("target_type", "VARCHAR(50)", False, None, True, None, "目标类型（多态）"),
            ("target_id", "BIGINT", False, None, True, None, "目标ID（多态）"),
            ("actor_id", "BIGINT", False, "users", True, None, "触发者ID（外键，可空）"),
            ("is_read", "BOOLEAN", False, None, False, "FALSE", "是否已读"),
            ("read_at", "TIMESTAMP", False, None, True, None, "阅读时间"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
            ("is_deleted", "BOOLEAN", False, None, False, "FALSE", "软删除标记"),
            ("deleted_at", "TIMESTAMP", False, None, True, None, "删除时间"),
        ],
    },
    {
        "name": "topic_collections",
        "cn_name": "专题合集表",
        "desc": "专题合集，聚合多条相关信息",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("title", "VARCHAR(200)", False, None, False, None, "专题标题"),
            ("description", "TEXT", False, None, True, None, "专题描述"),
            ("cover_url", "VARCHAR(500)", False, None, True, None, "封面URL"),
            ("school_id", "BIGINT", False, "schools", False, None, "学校ID（外键）"),
            ("creator_id", "BIGINT", False, "users", False, None, "创建者ID（外键）"),
            ("post_count", "INTEGER", False, None, False, "0", "信息数"),
            ("view_count", "INTEGER", False, None, False, "0", "浏览数"),
            ("status", "VARCHAR(20)", False, None, False, "'draft'", "状态"),
            ("sort_order", "INTEGER", False, None, False, "0", "排序"),
            ("published_at", "TIMESTAMP", False, None, True, None, "发布时间"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
            ("updated_at", "TIMESTAMP", False, None, False, "CURRENT", "更新时间"),
            ("is_deleted", "BOOLEAN", False, None, False, "FALSE", "软删除标记"),
            ("deleted_at", "TIMESTAMP", False, None, True, None, "删除时间"),
        ],
    },
    {
        "name": "topic_collection_posts",
        "cn_name": "专题信息关联表",
        "desc": "TopicCollection 与 Post 的多对多关联表",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("topic_collection_id", "BIGINT", False, "topic_collections", False, None, "专题ID（外键）"),
            ("post_id", "BIGINT", False, "posts", False, None, "信息ID（外键）"),
            ("sort_order", "INTEGER", False, None, False, "0", "排序"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
        ],
    },
    {
        "name": "drafts",
        "cn_name": "草稿表",
        "desc": "用户信息草稿，自动保存未发布内容",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("user_id", "BIGINT", False, "users", False, None, "用户ID（外键）"),
            ("title", "VARCHAR(200)", False, None, True, None, "草稿标题"),
            ("content", "TEXT", False, None, True, None, "草稿内容"),
            ("category_id", "BIGINT", False, "categories", True, None, "分类ID（外键，可空）"),
            ("location_id", "BIGINT", False, "locations", True, None, "地点ID（外键，可空）"),
            ("is_anonymous", "BOOLEAN", False, None, False, "FALSE", "是否匿名"),
            ("extra_data", "TEXT", False, None, True, None, "额外数据(JSON)"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "创建时间"),
            ("updated_at", "TIMESTAMP", False, None, False, "CURRENT", "更新时间"),
            ("is_deleted", "BOOLEAN", False, None, False, "FALSE", "软删除标记"),
            ("deleted_at", "TIMESTAMP", False, None, True, None, "删除时间"),
        ],
    },
    {
        "name": "browse_histories",
        "cn_name": "浏览历史表",
        "desc": "用户浏览信息的历史记录",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("user_id", "BIGINT", False, "users", False, None, "用户ID（外键）"),
            ("post_id", "BIGINT", False, "posts", False, None, "信息ID（外键）"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "浏览时间"),
        ],
    },
    {
        "name": "search_histories",
        "cn_name": "搜索历史表",
        "desc": "用户搜索关键词的历史记录",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("user_id", "BIGINT", False, "users", False, None, "用户ID（外键）"),
            ("keyword", "VARCHAR(200)", False, None, False, None, "搜索关键词"),
            ("result_count", "INTEGER", False, None, True, None, "结果数量"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "搜索时间"),
        ],
    },
    {
        "name": "admin_operation_logs",
        "cn_name": "管理员操作日志表",
        "desc": "管理员所有操作的审计日志",
        "fields": [
            ("id", "BIGINT", True, None, False, "AUTO", "主键，自增"),
            ("admin_id", "BIGINT", False, "users", False, None, "管理员ID（外键）"),
            ("action", "VARCHAR(50)", False, None, False, None, "操作类型"),
            ("target_type", "VARCHAR(50)", False, None, False, None, "目标类型"),
            ("target_id", "BIGINT", False, None, False, None, "目标ID"),
            ("detail", "TEXT", False, None, True, None, "操作详情"),
            ("ip_address", "VARCHAR(45)", False, None, True, None, "IP地址"),
            ("user_agent", "VARCHAR(500)", False, None, True, None, "User-Agent"),
            ("created_at", "TIMESTAMP", False, None, False, "CURRENT", "操作时间"),
        ],
    },
]

# =============================================================================
# 2. 关系定义（用于 ER 图）
# (from_table, from_cardinality, to_table, to_cardinality, relation_name)
# cardinality: "1" 或 "N" 或 "M"
# =============================================================================
RELATIONS = [
    ("schools", "1", "users", "N", "归属"),
    ("schools", "1", "posts", "N", "发布于"),
    ("schools", "1", "locations", "N", "拥有"),
    ("schools", "1", "topic_collections", "N", "属于"),
    ("users", "1", "posts", "N", "发布"),
    ("categories", "1", "posts", "N", "分类"),
    ("locations", "1", "posts", "N", "位于"),
    ("posts", "1", "post_images", "N", "含图"),
    ("posts", "1", "comments", "N", "评论"),
    ("posts", "1", "likes", "N", "点赞"),
    ("posts", "1", "favorites", "N", "收藏"),
    ("posts", "1", "validation_records", "N", "验证"),
    ("posts", "1", "reports", "N", "被举报"),
    ("posts", "1", "browse_histories", "N", "被浏览"),
    ("posts", "M", "topic_collections", "N", "收录于"),  # M:N 通过 topic_collection_posts
    ("users", "1", "comments", "N", "评论"),
    ("users", "1", "likes", "N", "点赞"),
    ("users", "1", "favorites", "N", "收藏"),
    ("users", "1", "validation_records", "N", "验证"),
    ("users", "1", "reports", "N", "举报"),  # reporter_id
    ("users", "1", "reports", "N", "处理"),  # handler_id
    ("users", "1", "notifications", "N", "接收"),
    ("users", "1", "notifications", "N", "触发"),  # actor_id
    ("users", "1", "drafts", "N", "草稿"),
    ("users", "1", "browse_histories", "N", "浏览"),
    ("users", "1", "search_histories", "N", "搜索"),
    ("users", "1", "admin_operation_logs", "N", "操作"),
    ("users", "1", "topic_collections", "N", "创建"),
    ("comments", "1", "comments", "N", "回复"),  # 自引用
    ("comments", "1", "reports", "N", "被举报"),
    ("categories", "1", "drafts", "N", "草稿分类"),
    ("locations", "1", "drafts", "N", "草稿地点"),
]

# 子系统分组（用于生成子系统 ER 图）
SUBSYSTEMS = {
    "用户子系统": {
        "tables": ["schools", "users", "drafts"],
        "desc": "用户、学校、草稿管理",
    },
    "信息子系统": {
        "tables": ["schools", "users", "categories",
                   "locations", "posts", "post_images",
                   "topic_collections", "topic_collection_posts"],
        "desc": "信息发布、分类、地点、专题",
    },
    "互动子系统": {
        "tables": ["posts", "users", "comments", "likes", "favorites",
                   "browse_histories", "search_histories"],
        "desc": "评论、点赞、收藏、浏览、搜索",
    },
    "治理子系统": {
        "tables": ["posts", "users", "validation_records", "reports",
                   "comments", "notifications"],
        "desc": "协同验证、举报、通知（核心创新）",
    },
    "管理子系统": {
        "tables": ["users", "posts", "comments", "admin_operation_logs"],
        "desc": "管理员操作日志审计",
    },
}


# =============================================================================
# 3. 生成 Excel
# =============================================================================

def generate_excel(output_path):
    """生成表结构 Excel 文件"""
    wb = Workbook()

    # 样式定义
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(name="微软雅黑", size=14, bold=True, color="2F5496")
    desc_font = Font(name="微软雅黑", size=10, italic=True, color="595959")
    pk_font = Font(name="微软雅黑", size=10, bold=True, color="C00000")
    fk_font = Font(name="微软雅黑", size=10, bold=True, color="2F5496")
    normal_font = Font(name="微软雅黑", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    pk_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fk_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")

    # --- Sheet 1: 总览 ---
    ws_overview = wb.active
    ws_overview.title = "总览"

    # 标题
    ws_overview.merge_cells("A1:F1")
    ws_overview["A1"] = "此刻校园 - 数据库表结构总览"
    ws_overview["A1"].font = Font(name="微软雅黑", size=16, bold=True, color="2F5496")
    ws_overview["A1"].alignment = center_align
    ws_overview.row_dimensions[1].height = 35

    ws_overview.merge_cells("A2:F2")
    ws_overview["A2"] = f"共 {len(TABLES)} 张表 | 数据库：openGauss 7.0.0-RC3 | 模拟核心：江南大学蠡湖校区"
    ws_overview["A2"].font = desc_font
    ws_overview["A2"].alignment = center_align

    # 表头
    headers = ["序号", "表名", "中文名", "字段数", "说明", "Sheet"]
    for col, header in enumerate(headers, 1):
        cell = ws_overview.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # 数据行
    for idx, table in enumerate(TABLES, 1):
        row = 4 + idx
        ws_overview.cell(row=row, column=1, value=idx).font = normal_font
        ws_overview.cell(row=row, column=2, value=table["name"]).font = Font(
            name="Consolas", size=10, bold=True
        )
        ws_overview.cell(row=row, column=3, value=table["cn_name"]).font = normal_font
        ws_overview.cell(row=row, column=4, value=len(table["fields"])).font = normal_font
        ws_overview.cell(row=row, column=5, value=table["desc"]).font = normal_font
        ws_overview.cell(row=row, column=6, value=table["name"]).font = normal_font

        for col in range(1, 7):
            ws_overview.cell(row=row, column=col).alignment = left_align
            ws_overview.cell(row=row, column=col).border = thin_border

    # 列宽
    ws_overview.column_dimensions["A"].width = 6
    ws_overview.column_dimensions["B"].width = 28
    ws_overview.column_dimensions["C"].width = 18
    ws_overview.column_dimensions["D"].width = 8
    ws_overview.column_dimensions["E"].width = 50
    ws_overview.column_dimensions["F"].width = 28

    # --- 每张表一个 Sheet ---
    for table in TABLES:
        ws = wb.create_sheet(title=table["name"])

        # 表标题
        ws.merge_cells("A1:G1")
        ws["A1"] = f"{table['name']}（{table['cn_name']}）"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        ws.row_dimensions[1].height = 30

        # 表说明
        ws.merge_cells("A2:G2")
        ws["A2"] = f"说明：{table['desc']}"
        ws["A2"].font = desc_font
        ws["A2"].alignment = left_align

        # 表头
        headers = ["序号", "字段名", "数据类型", "主键", "外键", "可空", "默认值", "说明"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[4].height = 25

        # 字段行
        for idx, (name, dtype, is_pk, fk_ref, nullable, default, desc) in enumerate(
            table["fields"], 1
        ):
            row = 4 + idx
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=dtype)
            ws.cell(row=row, column=4, value="PK" if is_pk else "")
            ws.cell(row=row, column=5, value=f"→{fk_ref}" if fk_ref else "")
            ws.cell(row=row, column=6, value="NULL" if nullable else "NOT NULL")
            ws.cell(row=row, column=7, value=default if default else "")
            ws.cell(row=row, column=8, value=desc)

            # 设置字体与填充
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.alignment = left_align
                cell.border = thin_border
                if col == 2:  # 字段名
                    cell.font = Font(name="Consolas", size=10, bold=True)
                elif col == 3:  # 数据类型
                    cell.font = Font(name="Consolas", size=10)
                else:
                    cell.font = normal_font

            # 主键行高亮
            if is_pk:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = pk_fill
                ws.cell(row=row, column=4).font = pk_font
            # 外键行高亮
            elif fk_ref:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = fk_fill
                ws.cell(row=row, column=5).font = fk_font

        # 列宽
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 6
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 45

        # 冻结表头
        ws.freeze_panes = "A5"

    wb.save(output_path)
    print(f"[OK] Excel 已生成：{output_path}")


# =============================================================================
# 4. 生成 SVG ER 图
# =============================================================================

class SvgERGenerator:
    """SVG ER 图生成器"""

    def __init__(self):
        self.entity_colors = {
            "core": ("#4472C4", "#DEEBF7"),       # 核心实体：蓝
            "config": ("#70AD47", "#E2EFDA"),     # 配置实体：绿
            "relation": ("#ED7D31", "#FCE4D6"),   # 关联实体：橙
            "interaction": ("#FFC000", "#FFF2CC"), # 互动实体：黄
            "governance": ("#C00000", "#FBE4E4"), # 治理实体：红
            "system": ("#7030A0", "#E4DFF0"),     # 系统实体：紫
        }

    def get_entity_category(self, table_name):
        """根据表名判断实体类别"""
        config_tables = {"schools", "categories", "locations"}
        relation_tables = {"post_images", "topic_collection_posts"}
        interaction_tables = {"comments", "likes", "favorites", "browse_histories", "search_histories", "drafts"}
        governance_tables = {"validation_records", "reports", "notifications"}
        system_tables = {"admin_operation_logs"}
        core_tables = {"users", "posts", "topic_collections"}

        if table_name in config_tables:
            return "config"
        elif table_name in relation_tables:
            return "relation"
        elif table_name in interaction_tables:
            return "interaction"
        elif table_name in governance_tables:
            return "governance"
        elif table_name in system_tables:
            return "system"
        else:
            return "core"

    def escape_xml(self, text):
        """转义 XML 特殊字符"""
        text = str(text)
        text = text.replace("&", "&amp;")
        text = text.replace("<", "&lt;")
        text = text.replace(">", "&gt;")
        text = text.replace('"', "&quot;")
        text = text.replace("'", "&apos;")
        return text

    def calculate_entity_box(self, table, show_all_fields=False):
        """计算实体框尺寸"""
        header_height = 28
        field_height = 16
        if show_all_fields:
            field_count = len(table["fields"])
        else:
            # 只显示 PK + FK + 关键字段
            field_count = sum(
                1 for f in table["fields"] if f[2] or f[3]
            )
            field_count = max(field_count, 3)  # 最少显示 3 行

        # 估算宽度
        max_field_len = max(
            len(f[0]) + len(f[1]) + 5 for f in table["fields"]
        ) if table["fields"] else 20
        width = max(180, min(260, max_field_len * 8))
        height = header_height + field_count * field_height + 10
        return width, height

    def generate_entity_svg(self, table, x, y, show_all_fields=False):
        """生成单个实体的 SVG"""
        category = self.get_entity_category(table["name"])
        border_color, fill_color = self.entity_colors[category]
        width, height = self.calculate_entity_box(table, show_all_fields)

        svg_parts = []
        # 实体框背景
        svg_parts.append(
            f'  <rect x="{x}" y="{y}" width="{width}" height="{height}" '
            f'rx="4" ry="4" fill="{fill_color}" stroke="{border_color}" stroke-width="2"/>'
        )
        # 标题栏背景
        svg_parts.append(
            f'  <rect x="{x}" y="{y}" width="{width}" height="28" '
            f'rx="4" ry="4" fill="{border_color}"/>'
        )
        # 标题栏底部修正（去掉圆角）
        svg_parts.append(
            f'  <rect x="{x}" y="{y+20}" width="{width}" height="8" fill="{border_color}"/>'
        )
        # 表名
        svg_parts.append(
            f'  <text x="{x + width // 2}" y="{y + 19}" text-anchor="middle" '
            f'font-family="微软雅黑" font-size="13" font-weight="bold" fill="white">'
            f'{self.escape_xml(table["name"])}</text>'
        )

        # 字段列表
        if show_all_fields:
            fields_to_show = table["fields"]
        else:
            fields_to_show = [f for f in table["fields"] if f[2] or f[3]]
            if len(fields_to_show) < 3:
                # 补充关键字段
                for f in table["fields"]:
                    if f not in fields_to_show and len(fields_to_show) < 4:
                        fields_to_show.append(f)

        field_y = y + 28 + 14
        for fname, ftype, is_pk, fk_ref, *_ in fields_to_show:
            prefix = ""
            if is_pk:
                prefix = "PK "
                font_weight = "bold"
                text_fill = "#C00000"
            elif fk_ref:
                prefix = "FK "
                font_weight = "bold"
                text_fill = "#2F5496"
            else:
                font_weight = "normal"
                text_fill = "#333333"

            # 截断过长的字段类型
            display_type = ftype if len(ftype) <= 16 else ftype[:14] + ".."

            svg_parts.append(
                f'  <text x="{x + 8}" y="{field_y}" font-family="Consolas" '
                f'font-size="11" font-weight="{font_weight}" fill="{text_fill}">'
                f'{self.escape_xml(prefix + fname)}</text>'
            )
            svg_parts.append(
                f'  <text x="{x + width - 8}" y="{field_y}" text-anchor="end" '
                f'font-family="Consolas" font-size="10" fill="#666666">'
                f'{self.escape_xml(display_type)}</text>'
            )
            # 分隔线
            svg_parts.append(
                f'  <line x1="{x + 4}" y1="{field_y + 3}" x2="{x + width - 4}" '
                f'y2="{field_y + 3}" stroke="#D9D9D9" stroke-width="0.5"/>'
            )
            field_y += 16

        return "\n".join(svg_parts), width, height

    def generate_relation_line(self, x1, y1, x2, y2, from_card, to_card, label=""):
        """生成关系连线（含基数标注）"""
        parts = []
        # 连线
        parts.append(
            f'  <line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" '
            f'stroke="#666666" stroke-width="1.5"/>'
        )
        # 起点基数
        parts.append(
            f'  <text x="{x1 + 8}" y="{y1 + 5}" font-family="Arial" '
            f'font-size="12" font-weight="bold" fill="#C00000">{from_card}</text>'
        )
        # 终点基数
        parts.append(
            f'  <text x="{x2 - 18}" y="{y2 + 5}" font-family="Arial" '
            f'font-size="12" font-weight="bold" fill="#C00000">{to_card}</text>'
        )
        # 关系标签
        if label:
            mid_x = (x1 + x2) // 2
            mid_y = (y1 + y2) // 2
            parts.append(
                f'  <rect x="{mid_x - len(label) * 6}" y="{mid_y - 10}" '
                f'width="{len(label) * 12}" height="16" rx="8" ry="8" '
                f'fill="white" stroke="#999999" stroke-width="0.5"/>'
            )
            parts.append(
                f'  <text x="{mid_x}" y="{mid_y + 3}" text-anchor="middle" '
                f'font-family="微软雅黑" font-size="10" fill="#333333">'
                f'{self.escape_xml(label)}</text>'
            )
        return "\n".join(parts)

    def generate_overview_er(self, output_path):
        """生成总体 ER 图（简化版，仅实体名 + 关系）"""
        # 网格布局：5 列 x 5 行
        cols = 5
        col_width = 240
        row_height = 100
        margin_x = 40
        margin_y = 60
        box_w = 180
        box_h = 36

        table_positions = {}
        for idx, table in enumerate(TABLES):
            row = idx // cols
            col = idx % cols
            x = margin_x + col * col_width
            y = margin_y + row * row_height
            table_positions[table["name"]] = (x, y)

        svg_width = margin_x * 2 + cols * col_width
        svg_height = margin_y * 2 + ((len(TABLES) - 1) // cols + 1) * row_height + 50

        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{svg_width}" height="{svg_height}" '
            f'viewBox="0 0 {svg_width} {svg_height}">',
            f'<style>text{{font-family:"微软雅黑";}}</style>',
            # 背景
            f'<rect width="{svg_width}" height="{svg_height}" fill="#FAFAFA"/>',
        ]

        # 标题
        parts.append(
            f'  <text x="{svg_width // 2}" y="30" text-anchor="middle" '
            f'font-size="20" font-weight="bold" fill="#2F5496">'
            f'此刻校园 - 总体 E-R 图（21 实体 + 35 联系）</text>'
        )
        parts.append(
            f'  <text x="{svg_width // 2}" y="50" text-anchor="middle" '
            f'font-size="12" fill="#595959">数据库：openGauss 7.0.0-RC3 | 模拟核心：江南大学蠡湖校区</text>'
        )

        # 先画关系线
        for from_t, from_c, to_t, to_c, label in RELATIONS:
            if from_t not in table_positions or to_t not in table_positions:
                continue
            x1, y1 = table_positions[from_t]
            x2, y2 = table_positions[to_t]
            # 连线从框中心到框中心
            cx1, cy1 = x1 + box_w // 2, y1 + box_h // 2
            cx2, cy2 = x2 + box_w // 2, y2 + box_h // 2
            parts.append(
                f'  <line x1="{cx1}" y1="{cy1}" x2="{cx2}" y2="{cy2}" '
                f'stroke="#BFBFBF" stroke-width="1" opacity="0.6"/>'
            )

        # 再画实体框（覆盖在连线上）
        for table in TABLES:
            x, y = table_positions[table["name"]]
            category = self.get_entity_category(table["name"])
            border_color, fill_color = self.entity_colors[category]

            parts.append(
                f'  <rect x="{x}" y="{y}" width="{box_w}" height="{box_h}" '
                f'rx="6" ry="6" fill="{fill_color}" stroke="{border_color}" stroke-width="2"/>'
            )
            parts.append(
                f'  <text x="{x + box_w // 2}" y="{y + 17}" text-anchor="middle" '
                f'font-size="13" font-weight="bold" fill="{border_color}">'
                f'{self.escape_xml(table["name"])}</text>'
            )
            parts.append(
                f'  <text x="{x + box_w // 2}" y="{y + 30}" text-anchor="middle" '
                f'font-size="10" fill="#595959">{self.escape_xml(table["cn_name"])}</text>'
            )

        # 图例
        legend_y = svg_height - 35
        legend_items = [
            ("核心实体", "core"),
            ("配置实体", "config"),
            ("关联实体", "relation"),
            ("互动实体", "interaction"),
            ("治理实体", "governance"),
            ("系统实体", "system"),
        ]
        legend_x = 40
        for label, cat in legend_items:
            border_c, fill_c = self.entity_colors[cat]
            parts.append(
                f'  <rect x="{legend_x}" y="{legend_y}" width="14" height="14" '
                f'rx="2" ry="2" fill="{fill_c}" stroke="{border_c}" stroke-width="1.5"/>'
            )
            parts.append(
                f'  <text x="{legend_x + 20}" y="{legend_y + 12}" '
                f'font-size="11" fill="#333333">{label}</text>'
            )
            legend_x += 110

        parts.append("</svg>")

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(parts))
        print(f"[OK] 总体 ER 图已生成：{output_path}")

    def generate_subsystem_er(self, subsystem_name, tables_list, desc, output_path):
        """生成子系统 ER 图（详细版，含字段）"""
        # 过滤出子系统涉及的表
        sub_tables = [t for t in TABLES if t["name"] in tables_list]
        if not sub_tables:
            return

        # 计算布局：根据表数量决定列数
        n = len(sub_tables)
        if n <= 4:
            cols = 2
        elif n <= 6:
            cols = 3
        else:
            cols = 3

        margin = 50
        col_gap = 60
        row_gap = 80

        # 先计算每个实体框的尺寸
        sizes = []
        for table in sub_tables:
            w, h = self.calculate_entity_box(table, show_all_fields=False)
            sizes.append((w, h))

        max_col_w = max(s[0] for s in sizes)
        max_row_h = max(s[1] for s in sizes)
        col_width = max_col_w + col_gap
        row_height = max_row_h + row_gap

        rows_count = (n + cols - 1) // cols
        svg_width = margin * 2 + cols * col_width
        svg_height = margin * 2 + rows_count * row_height + 60

        # 计算位置
        positions = {}
        for idx, table in enumerate(sub_tables):
            row = idx // cols
            col = idx % cols
            x = margin + col * col_width
            y = margin + 30 + row * row_height
            positions[table["name"]] = (x, y)

        parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{svg_width}" height="{svg_height}" '
            f'viewBox="0 0 {svg_width} {svg_height}">',
            f'<style>text{{font-family:"微软雅黑";}}</style>',
            f'<rect width="{svg_width}" height="{svg_height}" fill="#FAFAFA"/>',
        ]

        # 标题
        parts.append(
            f'  <text x="{svg_width // 2}" y="28" text-anchor="middle" '
            f'font-size="18" font-weight="bold" fill="#2F5496">'
            f'此刻校园 - {subsystem_name} E-R 图</text>'
        )
        parts.append(
            f'  <text x="{svg_width // 2}" y="48" text-anchor="middle" '
            f'font-size="11" fill="#595959">{desc}（{n} 实体）</text>'
        )

        # 画关系线（只画子系统内部的关系）
        for from_t, from_c, to_t, to_c, label in RELATIONS:
            if from_t not in positions or to_t not in positions:
                continue
            x1, y1 = positions[from_t]
            x2, y2 = positions[to_t]
            w1, h1 = sizes[sub_tables.index(next(t for t in sub_tables if t["name"] == from_t))]
            w2, h2 = sizes[sub_tables.index(next(t for t in sub_tables if t["name"] == to_t))]

            # 连线从框边缘到框边缘
            cx1 = x1 + w1 // 2
            cy1 = y1 + h1 // 2
            cx2 = x2 + w2 // 2
            cy2 = y2 + h2 // 2

            parts.append(
                f'  <line x1="{cx1}" y1="{cy1}" x2="{cx2}" y2="{cy2}" '
                f'stroke="#999999" stroke-width="1.5"/>'
            )

            # 基数标注
            dx = cx2 - cx1
            dy = cy2 - cy1
            length = (dx ** 2 + dy ** 2) ** 0.5
            if length > 0:
                # from 端
                fx = cx1 + dx * 25 / length
                fy = cy1 + dy * 25 / length
                # to 端
                tx = cx2 - dx * 25 / length
                ty = cy2 - dy * 25 / length
                parts.append(
                    f'  <text x="{fx}" y="{fy + 4}" text-anchor="middle" '
                    f'font-size="13" font-weight="bold" fill="#C00000">{from_c}</text>'
                )
                parts.append(
                    f'  <text x="{tx}" y="{ty + 4}" text-anchor="middle" '
                    f'font-size="13" font-weight="bold" fill="#C00000">{to_c}</text>'
                )
                # 关系标签
                if label:
                    mx = (cx1 + cx2) // 2
                    my = (cy1 + cy2) // 2
                    parts.append(
                        f'  <rect x="{mx - len(label) * 6 - 4}" y="{my - 9}" '
                        f'width="{len(label) * 12 + 8}" height="16" rx="8" ry="8" '
                        f'fill="white" stroke="#999999" stroke-width="0.5"/>'
                    )
                    parts.append(
                        f'  <text x="{mx}" y="{my + 3}" text-anchor="middle" '
                        f'font-size="10" fill="#333333">{self.escape_xml(label)}</text>'
                    )

        # 画实体框
        for table in sub_tables:
            x, y = positions[table["name"]]
            entity_svg, _, _ = self.generate_entity_svg(table, x, y, show_all_fields=False)
            parts.append(entity_svg)

        parts.append("</svg>")

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(parts))
        print(f"[OK] {subsystem_name} ER 图已生成：{output_path}")


# =============================================================================
# 5. 生成 DOT 源码（Graphviz）
# =============================================================================

def generate_dot(output_path):
    """生成 Graphviz DOT 源码"""
    lines = [
        'digraph ERDiagram {',
        '  rankdir=LR;',
        '  fontname="Microsoft YaHei";',
        '  fontsize=20;',
        '  label="此刻校园 - 数据库 E-R 图（21 实体 + 35 联系）";',
        '  labelloc="t";',
        '  bgcolor="#FAFAFA";',
        '  node [fontname="Consolas", fontsize=10];',
        '  edge [fontname="Microsoft YaHei", fontsize=9];',
        '',
        '  // ===== 实体节点定义 =====',
    ]

    category_colors = {
        "core": "#DEEBF7",
        "config": "#E2EFDA",
        "relation": "#FCE4D6",
        "interaction": "#FFF2CC",
        "governance": "#FBE4E4",
        "system": "#E4DFF0",
    }

    gen = SvgERGenerator()

    for table in TABLES:
        cat = gen.get_entity_category(table["name"])
        fillcolor = category_colors[cat]
        # 构建字段标签
        field_lines = []
        for fname, ftype, is_pk, fk_ref, *_ in table["fields"]:
            prefix = ""
            if is_pk:
                prefix = "PK "
            elif fk_ref:
                prefix = "FK "
            field_lines.append(f"    {prefix}{fname}: {ftype}")

        label = "{" + table["name"] + " (" + table["cn_name"] + ")\\l" + "\\l".join(field_lines) + "\\l}"
        label = label.replace('"', '\\"')

        lines.append(
            f'  {table["name"]} [shape=record, style=filled, fillcolor="{fillcolor}", '
            f'label="{label}"];'
        )

    lines.append('')
    lines.append('  // ===== 关系定义 =====')

    for from_t, from_c, to_t, to_c, label in RELATIONS:
        arrow = "arrowhead=crow" if to_c == "N" else "arrowhead=none"
        tail = "arrowtail=tee" if from_c == "1" else "arrowtail=crow"
        lines.append(
            f'  {from_t} -> {to_t} [{tail}, {arrow}, dir=both, '
            f'label="{label}"];'
        )

    lines.append('}')

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"[OK] DOT 源码已生成：{output_path}")


# =============================================================================
# 6. 主函数
# =============================================================================

def main():
    # 输出目录：项目根目录下的 docs/design/
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    output_dir = os.path.join(project_root, "docs", "design")
    os.makedirs(output_dir, exist_ok=True)

    print(f"输出目录：{output_dir}")
    print(f"表数量：{len(TABLES)}")
    print(f"关系数量：{len(RELATIONS)}")
    print("-" * 60)

    # 1. 生成 Excel
    excel_path = os.path.join(output_dir, "此刻校园_数据库表结构.xlsx")
    generate_excel(excel_path)

    # 2. 生成 ER 图（SVG）
    gen = SvgERGenerator()

    # 2.1 总体 ER 图
    overview_path = os.path.join(output_dir, "ER图_总体.svg")
    gen.generate_overview_er(overview_path)

    # 2.2 子系统 ER 图
    for sub_name, sub_info in SUBSYSTEMS.items():
        safe_name = sub_name.replace("子系统", "")
        sub_path = os.path.join(output_dir, f"ER图_{safe_name}子系统.svg")
        gen.generate_subsystem_er(
            sub_name, sub_info["tables"], sub_info["desc"], sub_path
        )

    # 3. 生成 DOT 源码
    dot_path = os.path.join(output_dir, "ER图_源码.dot")
    generate_dot(dot_path)

    print("-" * 60)
    print("全部生成完成！")
    print(f"输出目录：{output_dir}")
    # 列出所有生成的文件
    for f in sorted(os.listdir(output_dir)):
        fpath = os.path.join(output_dir, f)
        size = os.path.getsize(fpath)
        print(f"  {f} ({size:,} bytes)")


if __name__ == "__main__":
    main()
